import openpyxl
import json
from openpyxl.styles import Font, PatternFill

with open('clients.json', encoding='utf-8') as file:
    data = json.load(file)

book = openpyxl.Workbook() 
sheet = book.active

#доп описание
sheet['A2'] = 'ID'
sheet['B2'] = 'FIO'
sheet['C2'] = 'PHONE'
sheet['D2'] = 'CITY'
sheet['E2'] = 'EMAIL'

# Устанавливаем голубой цвет фона для заголовков
header_fill = PatternFill(start_color="00CCFFFF", end_color='00CCFFFF', fill_type='solid')
header_font = Font(bold=True)

# Применяем стили к заголовкам
for cell in sheet[2]:
    cell.fill = header_fill
    cell.font = header_font
    
row = 3
for client in data['clients']:
    sheet[row][0].value = client['id']
    sheet[row][1].value = ' '.join(client['fio'])
    sheet[row][2].value = ' '.join(client['phone'])
    sheet[row][3].value = client['city']
    sheet[row][4].value = client['email']
    row += 1

book.save('Test.xlsx')
book.close()
