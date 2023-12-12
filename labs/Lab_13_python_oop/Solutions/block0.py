import openpyxl
from openpyxl.styles import Font, PatternFill

def create_excel_file(filename, creation_time, start_date, end_date):
    book = openpyxl.Workbook()
    sheet = book.active

    # Добавляем метку времени создания в ячейку A1 98F5FF
    sheet['A1'] = 'Параметры запроса'
    sheet['A2'] = 'Дата выгрузки'
    sheet['B2'] = creation_time.strftime('%Y-%m-%d %H:%M:%S')
    sheet['A3'] = 'Период, за который сделана выгрузка'
    sheet['B3'] = f'{start_date.strftime("%Y-%m-%d")} - {end_date.strftime("%Y-%m-%d")}'
    
    header_fill = PatternFill(start_color="DEB887", end_color='DEB887', fill_type='solid')
    for cell in sheet[1]:
        cell.fill = header_fill

    header_fill1 = PatternFill(start_color="98F5FF", end_color='98F5FF', fill_type='solid')

    start_row = 2
    end_row = 3
    for row_num in range(start_row, end_row+1):
        row = sheet[row_num]
        for cell in row:
            cell.fill = header_fill1

    font = Font(bold=True)
    sheet['A2'].font = font
    sheet['A1'].font = font
    sheet['A3'].font = font

    book.save(filename)
    book.close()