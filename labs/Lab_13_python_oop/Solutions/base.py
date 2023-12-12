import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
# Чтение данных о клиентах из файла clients.json
with open('clients.json', encoding='utf-8') as clients_file:
    clients_data = json.load(clients_file)

clients = clients_data['clients']

# Чтение данных о платежах из файла payments.json
with open('payments.json', encoding='utf-8') as payments_file:
    payments_data = json.load(payments_file)

payments = payments_data['payments']

# Создание словаря для хранения баланса клиента
client_balances = {client['id']: 0 for client in clients}

# Обновление баланса клиента на основе платежей
for payment in payments:
    client_id = payment['client_id']
    amount = payment['amount']
    client_balances[client_id] += amount

# Сортировка клиентов по балансу
sorted_clients = sorted(clients, key=lambda client: client_balances[client['id']], reverse=True)

# Создание нового файла Excel
workbook = Workbook()
sheet = workbook.active

# Устанавливаем голубой цвет фона для заголовков
header_fill = PatternFill(start_color="00CCFFFF", end_color='00CCFFFF', fill_type='solid')
header_font = Font(bold=True)

for row in range(1, 4):
    for col in range(1, 4):
        cell = sheet.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font

for row in range(15, 18):
    for col in range(1, 4):
        cell = sheet.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font

# Запись данных о топ 10 задолжниках
sheet['A1'] = "Анализ состояния счёта"
sheet.merge_cells('B2:C2')
sheet['B2'] = "Задолженности"
sheet['B3'] = "Клиент"
sheet['C3'] = "Состояние счёта"
sheet.merge_cells('A2:A3')
sheet['A2'] = "Статистика состояния счёта клиента"

row = 4
top_debtors = [client for client in sorted_clients if client_balances[client['id']] < 0][-10:]
for client in reversed(top_debtors):
    sheet[f'A{row}'] = f"Клиент ID: {client['id']}"
    sheet[f'B{row}'] = f"ФИО: {client['fio']}"
    sheet[f'C{row}'] = f"Баланс: {client_balances[client['id']]}"
    row += 1

# Запись данных о топ 10 клиентах с самым большим балансом
row += 2
sheet[f'A{row}'] = "Статистика состояния счёта клиента"
sheet.merge_cells('B16:C16')
sheet['B16'] = "Прибыльность"
sheet['B17'] = "Клиент"
sheet['C17'] = "Состояние счёта"
sheet.merge_cells('A16:A17')

# Сортировка клиентов по балансу
sorted_clients = sorted(clients, key=lambda client: client_balances[client['id']])

row += 2
top_balances = [client for client in sorted_clients if client_balances[client['id']] >= 0][-10:]
for client in reversed(top_balances):
    sheet[f'A{row}'] = f"Клиент ID: {client['id']}"
    sheet[f'B{row}'] = f"ФИО: {client['fio']}"
    sheet[f'C{row}'] = f"Баланс: {client_balances[client['id']]}"
    row += 1

# Сохранение файла Excel
workbook.save(filename='Task4.xlsx')