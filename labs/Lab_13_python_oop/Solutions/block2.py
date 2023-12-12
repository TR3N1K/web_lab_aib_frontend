import openpyxl
import json
from openpyxl.styles import Font, PatternFill

with open('payments.json', encoding='utf-8') as file:
    data = json.load(file)

with open('clients.json', encoding='utf-8') as file:
    dataClient = json.load(file)

book = openpyxl.Workbook() 
sheet = book.active

#доп описание
sheet['A7'] = 'Топ клиентов по количеству платежей'
sheet['B7'] = 'Текущий квартал'
sheet['C7'] = 'Q2 2023 год'
sheet['D7'] = 'Q3 2023 год'
sheet['E7'] = 'Q4 2022 год'

# Устанавливаем голубой цвет фона для заголовков
header_fill = PatternFill(start_color="00CCFFFF", end_color='00CCFFFF', fill_type='solid')
header_font = Font(bold=True)

# Применяем стили к заголовкам
for cell in sheet[7]:
    cell.fill = header_fill
    cell.font = header_font

for item in file:
    if item == 403:
        print(item)

book.save('Test3.xlsx')
book.close()