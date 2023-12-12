import openpyxl
import json
from openpyxl.styles import Font, PatternFill
from openpyxl import Workbook
# Открываем файл clients.json и загружаем данные
with open('clients.json', encoding='utf-8') as file:
    data = json.load(file)

# Создаем словарь для подсчета количества клиентов в каждом городе
city_counts = {}

# Подсчитываем количество клиентов в каждом городе
for client in data['clients']:
    city = client['city']
    if city in city_counts:
        city_counts[city] += 1
    else:
        city_counts[city] = 1

# Сортируем города по количеству клиентов в порядке убывания
sorted_cities = sorted(city_counts.items(), key=lambda x: x[1], reverse=True)

# Ограничиваем список только топ-10 городами
top_10_cities = sorted_cities[:10]

# Создаем новую книгу Excel
wb = Workbook()

# Получаем активный лист
sheet = wb.active

# Устанавливаем голубой цвет фона для заголовков
header_fill = PatternFill(start_color="00CCFFFF", end_color='00CCFFFF', fill_type='solid')
header_font = Font(bold=True)

# Записываем заголовки столбцов
sheet['A1'] = 'Город'
sheet['B1'] = 'Количество клиентов'

for cell in sheet[1]:
    cell.fill = header_fill
    cell.font = header_font

# Записываем данные в лист
row = 2
for city, count in top_10_cities:
    sheet[f'A{row}'] = city
    sheet[f'B{row}'] = count
    row += 1

# Сохраняем файл
wb.save('Task3.xlsx')

print("Данные успешно записаны в файл Task3.xlsx!")